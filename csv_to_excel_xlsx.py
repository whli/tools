# coding=utf-8

import sys
reload(sys)
sys.setdefaultencoding("utf-8")

import xlwt
from openpyxl.workbook import Workbook
from openpyxl.writer.excel import ExcelWriter
from openpyxl.utils import get_column_letter

def print_help():
    print "USAGE: python csv_to_excel.py CSV_FILE EXCEL_FILE"
    return

def main():

    if len(sys.argv) != 3:
        print_help()
        return

    excel_file = Workbook()
    excel_writer = ExcelWriter(workbook=excel_file)
    sheet = excel_file.worksheets[0]

    csv_file = open(sys.argv[1], "r")

    row = 1
    for line in csv_file:
        line = line.strip().decode("utf-8", "ignore")
        ary = line.split("@")
        col = 1

        for item in ary:
            col_letter = get_column_letter(col)
            sheet.cell("%s%s" % (col_letter, row)).value = "%s" % (item.encode("utf-8", "ignore"))
            col = col + 1
        row = row + 1
        
    excel_writer.save(filename=sys.argv[2])

    return

main()

